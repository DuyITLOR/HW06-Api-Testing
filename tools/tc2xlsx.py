#!/usr/bin/env python3
"""tc2xlsx.py — gom các bảng test case Markdown thành 1 file Excel (§14 đòi 'Excel test cases and test summary').

    python3 tools/tc2xlsx.py            # -> excel/23127178_HW06_TestCases.xlsx

Nguồn (mỗi API 3 file, đúng 3 bước của §6):
    test-cases/<api-slug>/generated.md   bước 1 — test case do AI sinh
    test-cases/<api-slug>/audit.md       bước 2 — nhãn VALID/INVALID/INCOMPLETE + bản sửa
    test-cases/<api-slug>/extended.md    bước 3 — ≥5 case sinh viên tự thêm
cộng test-cases/test-summary/summary.md (do npm run summary sinh) và traceability-matrix.md.

Quy ước: script lấy bảng Markdown ĐẦU TIÊN trong file có cột đầu tên 'TC ID'. Viết bảng khác
thì sửa HEADER_HINT bên dưới, đừng sửa dữ liệu cho khớp script.
"""
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "excel" / "23127178_HW06_TestCases.xlsx"
HEADER_HINT = "tc id"

APIS = [
    ("api-01-products-search", "API-01 Pool A GET products"),
    ("api-02-cart-add", "API-02 Pool B POST cart"),
    ("api-03-product-update", "API-03 Pool C PUT product"),
]
STEPS = [("generated.md", "AI sinh"), ("audit.md", "Audit"), ("extended.md", "SV thêm")]

HDR_FILL = PatternFill("solid", fgColor="DCE6F1")
HDR_FONT = Font(bold=True)


def parse_tables(path: Path):
    """Trả về list bảng; mỗi bảng là (header, rows)."""
    if not path.exists():
        return []
    lines = path.read_text(encoding="utf-8").splitlines()
    tables, block = [], []
    for line in lines + [""]:
        if line.strip().startswith("|"):
            block.append(line.strip())
            continue
        if len(block) >= 2:
            cells = [[c.strip() for c in re.split(r"(?<!\\)\|", b)[1:-1]] for b in block]
            # dòng thứ hai của bảng Markdown là dòng ---
            if cells and all(set(c) <= set("-: ") for c in cells[1]):
                tables.append((cells[0], cells[2:]))
        block = []
    return tables


def pick(path: Path):
    for header, rows in parse_tables(path):
        if header and HEADER_HINT in header[0].lower():
            return header, rows
    tabs = parse_tables(path)
    return tabs[0] if tabs else (None, None)


def write_sheet(wb, title, header, rows, extra_col=None):
    ws = wb.create_sheet(title[:31])
    head = ([extra_col[0]] if extra_col else []) + header
    ws.append(head)
    for r in rows:
        ws.append(([extra_col[1]] if extra_col else []) + r)
    for i in range(1, len(head) + 1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [10] + [max(12, min(52, len(h) + 6)) for h in head[1:]] if extra_col else [max(12, min(52, len(h) + 6)) for h in head]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    total = 0
    index_rows = []

    for slug, label in APIS:
        # `audit.md` là **bản cuối** của cùng bộ case trong `generated.md` (đã dán nhãn + đã sửa), nên
        # sheet gộp lấy audit + extended. Gộp cả generated nữa là đếm đúp mỗi case AI hai lần —
        # lỗi này đã xảy ra ở lần chạy đầu (250 dòng thay vì 131).
        merged_header, merged_rows = None, []
        has_audit = (ROOT / "test-cases" / slug / "audit.md").exists()
        for fname, step in STEPS:
            header, rows = pick(ROOT / "test-cases" / slug / fname)
            if not header:
                index_rows.append([label, step, "chưa có", 0])
                continue
            index_rows.append([label, step, fname, len(rows)])
            if fname == "generated.md" and has_audit:
                continue  # đã có bản audit — không đưa vào sheet gộp
            if merged_header is None:
                merged_header = header
            if header != merged_header:
                print(f"  [LUU Y] {slug}/{fname} có cột khác — vẫn gộp, kiểm lại bằng tay")
            merged_rows += [["" if c is None else c for c in r] for r in rows]
        if merged_header:
            write_sheet(wb, label, merged_header, merged_rows, extra_col=None)
            total += len(merged_rows)
        else:
            print(f"  [BO QUA] {slug}: chưa có bảng test case nào")

    # Sheet tổng: đọc summary.md do Newman sinh, giữ nguyên bảng đầu tiên.
    s_header, s_rows = pick(ROOT / "test-cases" / "test-summary" / "summary.md")
    if s_header:
        write_sheet(wb, "Test Summary (Newman)", s_header, s_rows)
    t_header, t_rows = pick(ROOT / "test-cases" / "test-summary" / "traceability-matrix.md")
    if t_header:
        write_sheet(wb, "Traceability", t_header, t_rows)

    ws = wb.create_sheet("Index", 0)
    ws.append(["API", "Bước (§6)", "File nguồn", "Số test case"])
    for r in index_rows:
        ws.append(r)
    ws.append([])
    ws.append(["TỔNG test case", "", "", total])
    for i, w in enumerate([30, 14, 18, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill, c.font = HDR_FILL, HDR_FONT

    if total == 0:
        print("  [LOI] Chưa có test case nào để xuất. Viết test-cases/<api>/generated.md trước.")
        # Vẫn xuất file rỗng để thấy đúng cấu trúc, nhưng báo mã lỗi để không nhầm là đã xong.
        OUT.parent.mkdir(parents=True, exist_ok=True)
        wb.save(OUT)
        print(f"  → {OUT.relative_to(ROOT)} (rỗng)")
        return 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"  → {OUT.relative_to(ROOT)}  ·  {total} test case  ·  {len(wb.sheetnames)} sheet")
    return 0


if __name__ == "__main__":
    sys.exit(main())
